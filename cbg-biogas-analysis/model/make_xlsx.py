"""
make_xlsx.py — build ../model/cbg_model.xlsx with 4 sheets:
  Drivers     : every assumption, EDITABLE (yellow), with provenance + tier.
  LiveModel   : 15-yr cashflow in Excel FORMULAS referencing Drivers → IRR/NPV/DSCR
                recompute live when you edit a driver. Configured for M1 by default;
                set capex multiplier=3 + add logistics for M2; set revenue ₹/kg = avoided
                cost for M3. A Python cross-check column confirms the formulas match the engine.
  Scenarios   : Tailwind/Base/Stress snapshot (from the Python engine).
  Sensitivity : tornado table + Monte-Carlo P10/50/90 + frontier corners.
Run: python3 make_xlsx.py
"""
import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from drivers import Drivers
from cbg_model import project_financials, mass_energy_balance, weighted_yield_kg_per_t, FEEDSTOCKS
import scenarios as S

OUT = os.path.join(os.path.dirname(__file__), "cbg_model.xlsx")
d = Drivers()
EDIT = PatternFill("solid", fgColor="FFF2CC")     # editable
HDR  = PatternFill("solid", fgColor="4C78A8")
SUB  = PatternFill("solid", fgColor="D9E1F2")
BOLD = Font(bold=True); WHITE = Font(bold=True, color="FFFFFF")
THIN = Border(*[Side(style="thin", color="DDDDDD")]*4)

def style_header(ws, row, cols):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c); cell.fill = HDR; cell.font = WHITE; cell.alignment = Alignment(wrap_text=True, vertical="center")

wb = Workbook()

# ---------------- Sheet 1: Drivers ----------------
ws = wb.active; ws.title = "Drivers"
ws["A1"] = "CBG Techno-Economic Model — DRIVERS (yellow = editable)"; ws["A1"].font = Font(bold=True, size=13)
ws["A2"] = "Single source of truth. Edit yellow cells; LiveModel recomputes. See sources/citations.md for tiers."
hdr = ["#", "Driver", "Value", "Unit", "Tier", "Note"]; ws.append([]); ws.append(hdr); style_header(ws, 4, 6)
rows = [
 ("D6","plf", d.plf, "frac", "T2", "realized PLF yr2-3; silent killer"),
 ("D6","yr1_plf_ramp", d.yr1_plf_ramp, "frac", "INF", "year-1 ramp"),
 ("-","operating_basis_days", d.operating_basis_days, "days", "-", "PLF embeds downtime; 365×PLF"),
 ("D3","weighted_yield", round(weighted_yield_kg_per_t(d.mix),2), "kgCBG/t", "INF", "blend yield (dung 15, pressmud 40)"),
 ("D1","dung_price", FEEDSTOCKS['dung']['price_rs_per_t'], "₹/t", "T2", "NDDB ₹1/kg→spiral; >₹1500 kills viability"),
 ("D1","pressmud_price", FEEDSTOCKS['press_mud']['price_rs_per_t'], "₹/t", "T2", "₹500-600; SEASONAL Oct-Apr"),
 ("D7","dung_share", d.mix['dung'], "frac", "-", "co-digestion recipe"),
 ("D7","pressmud_share", d.mix['press_mud'], "frac", "-", "co-digestion recipe"),
 ("D2","collectable_dung_kg_per_animal_day", d.collectable_dung_kg_per_animal_day, "kg", "T2", "dairy-anchored 5; scattered 3"),
 ("D2","collection_radius_km_max", d.collection_radius_km_max, "km", "T2", "wet-dung economical radius"),
 ("D4","cbg_price_rs_per_kg_M1", d.cbg_price_rs_per_kg_M1, "₹/kg", "T2", "SATAT 85% of CNG, NO floor"),
 ("D4","cbg_realisation_rs_per_kg_M2", d.cbg_realisation_rs_per_kg_M2, "₹/kg", "INF", "bottled niche realisation"),
 ("D13","avoided_fuel_rs_per_kg_cbg_M3", d.avoided_fuel_rs_per_kg_cbg_M3, "₹/kg", "READER", "M3 avoided PNG/LPG/FO — PLUG REAL VALUE"),
 ("D8","fom_price_rs_per_t", d.fom_price_rs_per_t, "₹/t", "T1", "₹500-4500; base near MDA floor"),
 ("D8","fom_t_per_t_feedstock", d.fom_t_per_t_feedstock, "frac", "INF", "solid FOM fraction"),
 ("D8","fom_revenue_clears (1/0)", 1 if d.fom_revenue_clears else 0, "bool", "T2", "0 = realistic ZERO-FOM downside"),
 ("D5","capex_intensity_rs_per_kgday", d.capex_intensity_rs_per_kgday, "₹/(kg/d)", "INF", "₹40-50k @5TPD; small worse"),
 ("D5","bottling_capex_multiplier_M2", d.bottling_capex_multiplier_M2, "x", "T1", "M2 retail 2-5× injection"),
 ("D5","capex_multiplier_for_live (1=M1/M3, 3=M2)", 1.0, "x", "-", "set 3 to model M2 in LiveModel"),
 ("D10","subsidy_cap_pct_of_gross", d.subsidy_cap_pct_of_gross, "frac", "T2", "realistic 20-35%"),
 ("D10","include_dpi_subsidy (1/0)", 1 if d.include_dpi_subsidy else 0, "bool", "T1", "needs CGD GSA w/ 50% ToP"),
 ("-","mnre_cfa_rs_per_4800kgday", d.mnre_cfa_rs_per_4800kgday, "₹", "T1", "₹4 cr/4800 kgday"),
 ("-","mnre_cfa_cap_rs", d.mnre_cfa_cap_rs, "₹", "T1", "cap ₹10 cr"),
 ("-","dpi_pipeline_support_rs", d.dpi_pipeline_support_rs_M1, "₹", "T1", "~₹9.95 cr (M1 only)"),
 ("-","fixed_opex_pct_of_capex", d.fixed_opex_pct_of_capex, "frac", "INF", "O&M+manpower (PLF-independent)"),
 ("-","variable_opex_rs_per_kg", d.variable_opex_rs_per_kg, "₹/kg", "INF", "power/consumables"),
 ("-","logistics_opex_rs_per_kg (M2 only)", 0.0, "₹/kg", "INF", "set 15 for M2 bottled"),
 ("-","aggregation_capex_share", d.aggregation_capex_share, "frac", "T2", "reinvest yrs 5,10"),
 ("D11","debt_share", d.debt_share, "frac", "T2", "75:25 post-tightening"),
 ("D11","cost_of_debt", d.cost_of_debt, "frac", "INF", "~10.5% MCLR-linked"),
 ("-","debt_tenor_years", d.debt_tenor_years, "yrs", "T1", "10-12"),
 ("D11","cost_of_equity", d.cost_of_equity, "frac", "OPN", "reader hurdle ~20%, ESG haircut"),
 ("-","tax_rate", d.tax_rate, "frac", "T1", "25% corporate"),
 ("-","project_life_years", d.project_life_years, "yrs", "T1", "offtake tenor 15 yrs"),
 ("-","depreciation_years", d.depreciation_years, "yrs", "INF", "straight-line"),
 ("-","scale_tpd_cbg (LiveModel)", 5.0, "TPD", "-", "plant CBG output for LiveModel"),
]
r = 5
name_to_cell = {}
for (dn, name, val, unit, tier, note) in rows:
    ws.cell(row=r, column=1, value=dn)
    ws.cell(row=r, column=2, value=name)
    c = ws.cell(row=r, column=3, value=val); c.fill = EDIT; c.font = BOLD
    ws.cell(row=r, column=4, value=unit); ws.cell(row=r, column=5, value=tier); ws.cell(row=r, column=6, value=note)
    name_to_cell[name] = f"Drivers!$C${r}"
    r += 1
for col, w in zip("ABCDEF", [6, 40, 14, 10, 7, 46]):
    ws.column_dimensions[col].width = w
ws.freeze_panes = "A5"

def C(n): return name_to_cell[n]  # shortcut to a driver cell ref

# ---------------- Sheet 2: LiveModel ----------------
m = wb.create_sheet("LiveModel")
m["A1"] = "LIVE MODEL — edit Drivers, this recomputes (M1 default). Cross-check vs Python engine at right."
m["A1"].font = Font(bold=True, size=12)
# derived cells
m["A3"] = "cbg_nameplate_kgday"; m["B3"] = f"={C('scale_tpd_cbg (LiveModel)')}*1000"
m["A4"] = "feedstock_tpd";       m["B4"] = f"=B3/{C('weighted_yield')}"
m["A5"] = "blended_feed_₹/t";    m["B5"] = f"={C('dung_share')}*{C('dung_price')}+{C('pressmud_share')}*{C('pressmud_price')}"
m["A6"] = "gross_capex ₹";       m["B6"] = f"=B3*{C('capex_intensity_rs_per_kgday')}*{C('capex_multiplier_for_live (1=M1/M3, 3=M2)')}"
m["A7"] = "mnre_subsidy ₹";      m["B7"] = f"=MIN({C('mnre_cfa_rs_per_4800kgday')}*(B3/4800),{C('mnre_cfa_cap_rs')})"
m["A8"] = "dpi_subsidy ₹";       m["B8"] = f"=IF({C('include_dpi_subsidy (1/0)')}=1,{C('dpi_pipeline_support_rs')},0)"
m["A9"] = "subsidy ₹";           m["B9"] = f"=MIN(B7+B8,{C('subsidy_cap_pct_of_gross')}*B6)"
m["A10"] = "net_capex ₹";        m["B10"] = "=B6-B9"
m["A11"] = "debt ₹";             m["B11"] = f"=B10*{C('debt_share')}"
m["A12"] = "equity ₹";           m["B12"] = f"=B10*(1-{C('debt_share')})"
m["A13"] = "depreciation ₹/yr";  m["B13"] = f"=B10/{C('depreciation_years')}"
m["A14"] = "WACC";               m["B14"] = f"={C('debt_share')}*{C('cost_of_debt')}*(1-{C('tax_rate')})+(1-{C('debt_share')})*{C('cost_of_equity')}"
m["A15"] = "revenue_unit ₹/kg (M1)"; m["B15"] = f"={C('cbg_price_rs_per_kg_M1')}"; m["B15"].fill = EDIT
for cell in ["B3","B4","B5","B6","B7","B8","B9","B10","B11","B12","B13","B14"]:
    m[cell].number_format = "#,##0"
m["B14"].number_format = "0.0%"

# cashflow table (rows 18..34 = year 0..15+header)
hrow = 18
heads = ["Year","PLF","CBG kg/yr","Feed t/yr","Feedstock ₹","Fixed opex ₹","Var opex ₹","Logistics ₹",
         "Revenue ₹","EBIT ₹","Reinvest ₹","Unlev FCF ₹","Debt bal ₹","Interest ₹","Principal ₹","Equity CF ₹","DSCR"]
for j, h in enumerate(heads, start=1):
    cc = m.cell(row=hrow, column=j, value=h); cc.fill = HDR; cc.font = WHITE; cc.alignment = Alignment(wrap_text=True)
# year 0
m.cell(row=hrow+1, column=1, value=0)
m.cell(row=hrow+1, column=12, value="=-B10")     # unlev FCF yr0
m.cell(row=hrow+1, column=13, value="=B11")      # debt balance start
m.cell(row=hrow+1, column=16, value="=-B12")     # equity CF yr0
for y in range(1, 16):
    rr = hrow + 1 + y
    plf = f'=IF(A{rr}=1,{C("yr1_plf_ramp")},{C("plf")})'
    m.cell(row=rr, column=1, value=y)
    m.cell(row=rr, column=2, value=plf)
    m.cell(row=rr, column=3, value=f"=B$3*{C('operating_basis_days')}*B{rr}")
    m.cell(row=rr, column=4, value=f"=B$4*{C('operating_basis_days')}*B{rr}")
    m.cell(row=rr, column=5, value=f"=D{rr}*B$5")
    m.cell(row=rr, column=6, value=f"={C('fixed_opex_pct_of_capex')}*B$6")
    m.cell(row=rr, column=7, value=f"=C{rr}*{C('variable_opex_rs_per_kg')}")
    m.cell(row=rr, column=8, value=f"=C{rr}*{C('logistics_opex_rs_per_kg (M2 only)')}")
    m.cell(row=rr, column=9, value=f"=C{rr}*B$15+IF({C('fom_revenue_clears (1/0)')}=1,D{rr}*{C('fom_t_per_t_feedstock')}*{C('fom_price_rs_per_t')},0)")
    m.cell(row=rr, column=10, value=f"=I{rr}-E{rr}-F{rr}-G{rr}-H{rr}-B$13")  # EBIT
    m.cell(row=rr, column=11, value=f"=IF(OR(A{rr}=5,A{rr}=10),{C('aggregation_capex_share')}*B$6,0)")
    m.cell(row=rr, column=12, value=f"=J{rr}-IF(J{rr}>0,J{rr}*{C('tax_rate')},0)+B$13-K{rr}")  # unlev FCF
    # debt schedule
    m.cell(row=rr, column=13, value=f"=M{rr-1}-O{rr-1}" if y > 1 else "=B11")
    m.cell(row=rr, column=14, value=f"=M{rr}*{C('cost_of_debt')}")
    m.cell(row=rr, column=15, value=f"=IF(A{rr}<={C('debt_tenor_years')},B$11/{C('debt_tenor_years')},0)")
    # equity CF = (EBIT-interest) - tax + dep - principal - reinvest
    m.cell(row=rr, column=16, value=f"=(J{rr}-N{rr})-IF((J{rr}-N{rr})>0,(J{rr}-N{rr})*{C('tax_rate')},0)+B$13-O{rr}-K{rr}")
    m.cell(row=rr, column=17, value=f'=IF((N{rr}+O{rr})>0,((J{rr}-N{rr})-IF((J{rr}-N{rr})>0,(J{rr}-N{rr})*{C("tax_rate")},0)+B$13+N{rr})/(N{rr}+O{rr}),"")')
# fix yr0 debt balance label collision: set M(hrow+1) explicitly already done
# results block
res = hrow + 18
m.cell(row=res, column=1, value="RESULTS").font = BOLD
lo, hi = hrow+1, hrow+16
m.cell(row=res+1, column=1, value="Project IRR"); m.cell(row=res+1, column=2, value=f"=IRR(L{lo}:L{hi})").number_format="0.0%"
m.cell(row=res+2, column=1, value="Equity IRR");  m.cell(row=res+2, column=2, value=f"=IRR(P{lo}:P{hi})").number_format="0.0%"
m.cell(row=res+3, column=1, value="NPV @ WACC ₹"); m.cell(row=res+3, column=2, value=f"=L{lo}+NPV(B14,L{lo+1}:L{hi})").number_format="#,##0"
m.cell(row=res+4, column=1, value="Min DSCR");     m.cell(row=res+4, column=2, value=f"=MIN(Q{lo+1}:Q{hi})").number_format="0.00"
# Python cross-check
base = project_financials(5.0, "M1_injection", d)
m.cell(row=res+1, column=4, value="Python engine →"); m.cell(row=res+1, column=5, value=round(base['proj_irr'],4)).number_format="0.0%"
m.cell(row=res+2, column=5, value=round(base['eq_irr'],4)).number_format="0.0%"
m.cell(row=res+3, column=5, value=round(base['npv'],0)).number_format="#,##0"
m.cell(row=res+4, column=5, value=round(base['min_dscr'],2)).number_format="0.00"
m.cell(row=res, column=4, value="(should match column B at base drivers)").font = Font(italic=True)
for col, w in zip("ABCDEFGHIJKLMNOPQ", [16,7,12,11,12,12,11,11,13,12,11,13,13,11,11,13,7]):
    m.column_dimensions[col].width = w

# ---------------- Sheet 3: Scenarios ----------------
sc = wb.create_sheet("Scenarios")
sc["A1"] = "SCENARIOS — M1 injection @ 5 TPD (Python engine snapshot)"; sc["A1"].font = Font(bold=True, size=12)
sc.append([]); sc.append(["Scenario","Project IRR","Equity IRR","NPV @WACC ₹cr","Min DSCR","Payback (yr)"]); style_header(sc, 3, 6)
for name, rr in S.run_scenarios().items():
    sc.append([name, rr['proj_irr'], rr['eq_irr'], round(rr['npv']/1e7,2),
               round(rr['min_dscr'],2) if rr['min_dscr'] else None, rr['payback'] or ">15"])
for row in sc.iter_rows(min_row=4, max_row=6, min_col=2, max_col=3):
    for c in row: c.number_format = "0.0%"
sc["A8"] = ("Tailwind = PLF .70 + DPI subsidy + FOM clears @₹1500 + CBG ₹85 + dung eased. "
            "Base = PLF .50, no DPI, FOM ₹1250, CBG ₹77, dung ₹1250. "
            "Stress = PLF .45, feedstock +40%, FOM=0, CBG ₹70 (OMC gaps).")
sc["A8"].alignment = Alignment(wrap_text=True); sc.merge_cells("A8:F11")
for col, w in zip("ABCDEF", [12,13,12,14,10,12]): sc.column_dimensions[col].width = w

# ---------------- Sheet 4: Sensitivity ----------------
se = wb.create_sheet("Sensitivity")
se["A1"] = "SENSITIVITY — tornado, Monte-Carlo, frontier (Python engine)"; se["A1"].font = Font(bold=True, size=12)
se.append([]); se.append(["TORNADO (equity IRR, M1@5TPD)","low","high","swing (pts)"]); style_header(se, 3, 4)
base_irr, trows = S.tornado()
for label, lo_, hi_, sw in trows:
    se.append([label, lo_, hi_, round(sw*100,1)])
for row in se.iter_rows(min_row=4, max_row=3+len(trows), min_col=2, max_col=3):
    for c in row: c.number_format = "0.0%"
rbase = 3 + len(trows) + 2
se.cell(row=rbase, column=1, value=f"Base equity IRR = {base_irr*100:.1f}%").font = BOLD
mc = S.monte_carlo()
se.cell(row=rbase+2, column=1, value="MONTE-CARLO (5,000 draws, equity IRR)").font = BOLD
for i,(k,v) in enumerate([("P10",mc['p10']),("P50",mc['p50']),("P90",mc['p90']),
                          ("P(IRR>WACC)",mc['prob_gt_wacc']),("P(IRR>20%)",mc['prob_gt_20']),
                          ("P(IRR<0)",mc['prob_negative'])]):
    se.cell(row=rbase+3+i, column=1, value=k); cc=se.cell(row=rbase+3+i, column=2, value=round(v,4)); cc.number_format="0.0%"
dungs,cbgs,grid,wacc = S.frontier()
se.cell(row=rbase+10, column=1, value="FRONTIER project IRR — best/worst corner").font = BOLD
se.cell(row=rbase+11, column=1, value=f"dung ₹{dungs[0]:.0f} / CBG ₹{cbgs[-1]:.0f}"); se.cell(row=rbase+11, column=2, value=round(grid[-1][0],4)).number_format="0.0%"
se.cell(row=rbase+12, column=1, value=f"dung ₹{dungs[-1]:.0f} / CBG ₹{cbgs[0]:.0f}"); se.cell(row=rbase+12, column=2, value=round(grid[0][-1],4)).number_format="0.0%"
for col, w in zip("ABCD", [34,12,12,12]): se.column_dimensions[col].width = w

wb.save(OUT)
print("wrote", OUT)
